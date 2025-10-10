import os
import re
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gradio as gr
import gradio.themes as gthemes

# ============================================================
# 🧩 Normalization helpers & constants
# ============================================================

SUFFIXES = {
    "ltd","limited","co","company","corp","corporation","inc","incorporated",
    "plc","public","llc","lp","llp","ulc","pc","pllc","sa","ag","nv","se","bv",
    "oy","ab","aps","as","kft","zrt","rt","sarl","sas","spa","gmbh","ug","bvba",
    "cvba","nvsa","pte","pty","bhd","sdn","kabushiki","kaisha","kk","godo","dk",
    "dmcc","pjsc","psc","jsc","ltda","srl","s.r.l","group","holdings","limitedpartnership"
}

COUNTRY_EQUIVALENTS = {
    "uk":"united kingdom","u.k.":"united kingdom","england":"united kingdom",
    "great britain":"united kingdom","britain":"united kingdom",
    "usa":"united states","u.s.a.":"united states","us":"united states",
    "america":"united states","united states of america":"united states",
    "uae":"united arab emirates","u.a.e.":"united arab emirates",
    "south korea":"republic of korea","korea":"republic of korea",
    "north korea":"democratic people's republic of korea","russia":"russian federation",
    "czechia":"czech republic","côte d’ivoire":"ivory coast","cote d'ivoire":"ivory coast",
    "iran":"islamic republic of iran","venezuela":"bolivarian republic of venezuela",
    "taiwan":"republic of china","hong kong sar":"hong kong","macao sar":"macau","prc":"china"
}

# ============================================================
# 🧹 Cleaning helpers
# ============================================================

def _normalize_tokens(text: str) -> str:
    if not isinstance(text, str):
        return ""
    text = re.sub(r"[^a-zA-Z0-9\s]", " ", text.lower())
    parts = [w for w in text.split() if w not in SUFFIXES]
    return " ".join(parts).strip()

def _clean_domain(domain: str) -> str:
    if not isinstance(domain, str):
        return ""
    domain = domain.lower().strip()
    domain = re.sub(r"^https?://", "", domain)
    domain = re.sub(r"/.*$", "", domain)
    domain = re.sub(r"^www\.", "", domain)
    parts = domain.split(".")
    return parts[-2] if len(parts) >= 2 else domain

def _extract_domain_from_email(email: str) -> str:
    if not isinstance(email, str) or "@" not in email:
        return ""
    domain = email.split("@")[-1].lower().strip()
    domain = re.sub(r"^www\.", "", domain)
    domain = re.sub(r"/.*$", "", domain)
    return domain

# ============================================================
# 🌐 Advanced Company ↔ Domain Comparison
# ============================================================

def compare_company_domain(company: str, domain: str):
    """Advanced company ↔ domain comparison with dynamic short-name & alias matching."""
    if not isinstance(company, str) or not isinstance(domain, str):
        return "Unsure – Please Check", 0, "missing input"

    c = _normalize_tokens(company)
    d_raw = domain.lower().strip()
    d = _clean_domain(d_raw)

    # --- Known brand aliases ---
    aliases = {
        "johnlewis": "john lewis group",
        "dlg": "direct line group",
        "thg": "the hut group",
        "ihg": "intercontinental hotels group",
        "bt": "british telecom",
        "hsbc": "hong kong and shanghai banking corporation",
        "pwc": "pricewaterhousecoopers",
        "kpmg": "kpmg international",
        "ey": "ernst and young",
        "dhl": "deutsche post",
        "unilever": "unilever plc",
        "tesco": "tesco plc",
    }

    if d in aliases:
        d = aliases[d]

    # --- Dynamic short domain detection (2–4 chars) ---
    if len(d) <= 4:
        letters_in_name = sum([1 for ch in d if ch in c])
        if letters_in_name >= len(d) - 1:
            return "Likely Match", 90, f"short domain pattern ({d}) found in name"

    # --- Direct containment ---
    if d.replace(" ", "") in c.replace(" ", "") or c.replace(" ", "") in d.replace(" ", ""):
        return "Likely Match", 100, "direct containment"

    # --- Shared industry term boost ---
    common_terms = {"energy", "power", "pharma", "bank", "group", "telecom", "systems", "media", "health", "tech"}
    if any(t in c.split() for t in common_terms if t in d):
        if fuzz.partial_ratio(c, d) >= 70:
            return "Likely Match", 85, "shared industry term"

    # --- Fuzzy matching fallback ---
    score_token = fuzz.token_set_ratio(c, d)
    score_partial = fuzz.partial_ratio(c, d)
    score_avg = (score_token + score_partial) / 2

    if score_avg >= 80:
        return "Likely Match", score_avg, "strong fuzzy"
    elif score_avg >= 65:
        return "Unsure – Please Check", score_avg, "moderate fuzzy"
    else:
        return "Likely NOT Match", score_avg, "low similarity"

# ============================================================
# 🧮 Main matching function
# ============================================================

def run_matching(master_file, picklist_file, highlight_changes=True, progress=gr.Progress(track_tqdm=True)):
    try:
        progress(0, desc="📂 Reading uploaded files...")
        df_master = pd.read_excel(master_file.name)
        df_picklist = pd.read_excel(picklist_file.name)

        progress(0.2, desc="🔧 Preparing data...")
        EXACT_PAIRS = [
            ("c_industry","c_industry"),
            ("asset_title","asset_title"),
            ("lead_country","lead_country"),
            ("departments","departments"),
            ("c_state","c_state")
        ]
        df_out = df_master.copy()
        corrected_cells = set()

        progress(0.4, desc="🔍 Matching Master ↔ Picklist...")
        for master_col, picklist_col in EXACT_PAIRS:
            out_col = f"Match_{master_col}"
            if master_col in df_master.columns and picklist_col in df_picklist.columns:
                pick_map = {v.strip().lower(): v.strip() for v in df_picklist[picklist_col].dropna().astype(str)}
                matches, new_vals = [], []
                for i, val in enumerate(df_master[master_col].fillna("").astype(str)):
                    val_norm = val.strip().lower()
                    val_norm_eq = COUNTRY_EQUIVALENTS.get(val_norm, val_norm) if master_col.lower() in ["lead_country","country","c_country"] else val_norm
                    if val_norm_eq in pick_map:
                        matches.append("Yes")
                        new_val = pick_map[val_norm_eq]
                        new_vals.append(new_val)
                        if new_val != val:
                            corrected_cells.add((master_col, i + 2))
                    else:
                        matches.append("No")
                        new_vals.append(val)
                df_out[out_col] = matches
                df_out[master_col] = new_vals
            else:
                df_out[out_col] = "Column Missing"

        # --- Dynamic question columns ---
        q_cols = [c for c in df_picklist.columns if re.match(r"(?i)q0*\d+|question\s*\d+", c)]
        for qc in q_cols:
            out_col = f"Match_{qc}"
            if qc in df_master.columns:
                valid_answers = set(df_picklist[qc].dropna().astype(str).str.strip().str.lower())
                matches = []
                for val in df_master[qc].fillna("").astype(str):
                    val_norm = val.strip().lower()
                    if val_norm in valid_answers:
                        matches.append("Yes")
                    elif val_norm == "":
                        matches.append("Blank")
                    else:
                        matches.append("No")
                df_out[out_col] = matches
            else:
                df_out[out_col] = "Column Missing"

        # --- Seniority parsing ---
        def parse_seniority(title):
            if not isinstance(title, str): return "Entry", "no title"
            t = title.lower().strip()
            if re.search(r"\bchief\b|\bcio\b|\bcto\b|\bceo\b|\bcfo\b|\bciso\b|\bcpo\b|\bcso\b|\bcoo\b|\bchro\b|\bpresident\b", t): return "C Suite", "c-level"
            if re.search(r"\bvice president\b|\bvp\b|\bsvp\b", t): return "VP", "vp"
            if re.search(r"\bhead\b", t): return "Head", "head"
            if re.search(r"\bdirector\b", t): return "Director", "director"
            if re.search(r"\bmanager\b|\bmgr\b", t): return "Manager", "manager"
            if re.search(r"\bsenior\b|\bsr\b|\blead\b|\bprincipal\b", t): return "Senior", "senior"
            if re.search(r"\bintern\b|\btrainee\b|\bassistant\b|\bgraduate\b", t): return "Entry", "entry"
            return "Entry", "none"

        if "jobtitle" in df_master.columns:
            parsed = df_master["jobtitle"].apply(parse_seniority)
            df_out["Parsed_Seniority"] = parsed.apply(lambda x: x[0])
            df_out["Seniority_Logic"] = parsed.apply(lambda x: x[1])
        else:
            df_out["Parsed_Seniority"] = None
            df_out["Seniority_Logic"] = "jobtitle column not found"

        # --- Domain vs Company (always from email) ---
        progress(0.6, desc="🌐 Validating company ↔ email domain...")
        company_cols = [c for c in df_master.columns if c.strip().lower() in ["companyname","company","company name","company_name"]]
        email_cols = [c for c in df_master.columns if "email" in c.lower()]

        if company_cols and email_cols:
            company_col = company_cols[0]
            email_col = email_cols[0]

            statuses, scores, reasons, email_domains = [], [], [], []

            for i in range(len(df_master)):
                comp = df_master.at[i, company_col]
                email_val = df_master.at[i, email_col]
                dom = _extract_domain_from_email(email_val)
                status, score, reason = compare_company_domain(comp, dom)

                statuses.append(status)
                scores.append(score)
                reasons.append(reason)
                email_domains.append(dom)

            df_out["Extracted_Email_Domain"] = email_domains
            df_out["Domain_Check_Status"] = statuses
            df_out["Domain_Check_Score"] = scores
            df_out["Domain_Check_Reason"] = reasons
        else:
            df_out["Domain_Check_Status"] = "Missing company or email column"
            df_out["Domain_Check_Score"] = None
            df_out["Domain_Check_Reason"] = None

        # --- Save + formatting ---
        progress(0.9, desc="💾 Saving results...")
        out_file = f"{os.path.splitext(master_file.name)[0]} - Full_Check_Results.xlsx"
        df_out.to_excel(out_file, index=False)
        wb = load_workbook(out_file)
        ws = wb.active
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        for col_idx, col in enumerate(df_out.columns, start=1):
            if col.startswith("Match_"):
                for row in range(2, ws.max_row + 1):
                    val = str(ws.cell(row=row, column=col_idx).value).strip().lower()
                    if val == "yes":
                        ws.cell(row=row, column=col_idx).fill = green
                    elif val == "no":
                        ws.cell(row=row, column=col_idx).fill = red
                    else:
                        ws.cell(row=row, column=col_idx).fill = yellow

        if highlight_changes:
            for col_name, row in corrected_cells:
                if col_name in df_out.columns:
                    col_idx = list(df_out.columns).index(col_name) + 1
                    ws.cell(row=row, column=col_idx).fill = blue

        wb.save(out_file)
        progress(1.0, desc="✅ Done! File ready for download.")
        return out_file

    except Exception as e:
        return f"❌ Error: {str(e)}"

# ============================================================
# 🎨 Fancy UI Theme
# ============================================================

fancy_theme = gthemes.Soft(
    primary_hue="blue",
    secondary_hue="indigo",
    neutral_hue="slate",
    text_size="md",
    radius_size="lg",
).set(
    font="Poppins",
    body_background_fill="#f8fafc",
    block_background_fill="#ffffff",
    border_color_primary="#d1d5db",
    button_primary_background_fill="#2563eb",
    button_primary_text_color="white",
)

custom_css = """
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;500;600&display=swap');
body, .gradio-container {
  font-family: 'Poppins', sans-serif !important;
  background: linear-gradient(180deg, #f9fafb 0%, #eef2ff 100%) !important;
}
h1, h2, h3, .title {
  color: #1e293b !important;
  font-weight: 600 !important;
}
.gr-button {
  background: linear-gradient(90deg, #2563eb, #4f46e5) !important;
  color: white !important;
  border-radius: 12px !important;
  font-weight: 600 !important;
  transition: all 0.2s ease-in-out !important;
}
.gr-button:hover {
  transform: scale(1.05);
  box-shadow: 0 4px 14px rgba(37,99,235,0.3);
}
footer {
  text-align: center;
  font-size: 13px;
  color: #64748b;
  margin-top: 20px;
}
"""

# ============================================================
# 🎛️ Gradio Interface
# ============================================================

demo = gr.Interface(
    fn=run_matching,
    inputs=[
        gr.File(label="📂 Upload MASTER Excel file (.xlsx)"),
        gr.File(label="📋 Upload PICKLIST Excel file (.xlsx)"),
        gr.Checkbox(label="Highlight changed values (💙 Blue Cells)", value=True)
    ],
    outputs=gr.File(label="⬇️ Download Processed Results"),
    title="✨ Master–Picklist + Email Domain Matching Tool",
    description=(
        "<div style='font-size:16px; line-height:1.6; color:#334155;'>"
        "<b>Welcome!</b> Upload your <span style='color:#2563eb;'>MASTER</span> "
        "and <span style='color:#4f46e5;'>PICKLIST</span> Excel files to automatically "
        "match company data, validate <b>email domains</b>, and highlight differences.<br><br>"
        "✅ Always uses domains extracted from emails<br>"
        "✅ Smart fuzzy + alias + short-domain matching<br>"
        "✅ Country normalization & color-coded results"
        "</div>"
    ),
    theme=fancy_theme,
    css=custom_css,
)

# ============================================================
# 🚀 Launch (Railway compatible)
# ============================================================

# ============================================================
# 🚀 Launch (Railway compatible)
# ============================================================

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 7860))
    demo.queue(concurrency_count=1, max_size=10).launch(
        server_name="0.0.0.0",     # required for Railway external access
        server_port=port,          # binds to dynamic port
        share=False,               # disable gradio.live
        show_api=False,            # suppress API display
        quiet=True,                # cleaner logs
        favicon_path=None
    )

