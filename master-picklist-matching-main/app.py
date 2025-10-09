import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gradio as gr

# ------------------------------------------------------------
# Core matching function
# ------------------------------------------------------------
def run_matching(master_file, picklist_file, progress=gr.Progress(track_tqdm=True)):
    try:
        progress(0, desc="Reading uploaded files...")
        df_master = pd.read_excel(master_file.name)
        df_picklist_raw = pd.read_excel(picklist_file.name)

        progress(0.2, desc="Preparing data...")

        EXACT_PAIRS = [
            ("c_industry", "c_industry"),
            ("asset_title", "asset_title"),
            ("lead_country", "lead_country"),
            ("departments", "departments"),
            ("c_state", "c_state"),
        ]

        keep_cols = ["c_industry", "asset_title", "lead_country", "departments", "c_state", "seniority"]
        df_picklist = df_picklist_raw[[c for c in keep_cols if c in df_picklist_raw.columns]].dropna(how="all").reset_index(drop=True)

        def normalize(s):
            return s.fillna("").astype(str).str.strip().str.lower()

        df_out = df_master.copy()
        corrected_cells = set()

        progress(0.4, desc="Running matching logic...")

        for master_col, picklist_col in EXACT_PAIRS:
            out_col = f"Match_{master_col}"
            if master_col in df_master.columns and picklist_col in df_picklist.columns:
                pick_map = {v.strip().lower(): v.strip() for v in df_picklist[picklist_col].dropna().astype(str)}
                matches, new_vals = [], []
                for i, val in enumerate(df_master[master_col].fillna("").astype(str)):
                    val_norm = val.strip().lower()
                    if val_norm in pick_map:
                        matches.append("Yes")
                        new_val = pick_map[val_norm]
                        new_vals.append(new_val)
                        if new_val != val:
                            corrected_cells.add((master_col, i + 2))
                    else:
                        matches.append("No")
                        new_vals.append(val)
                df_out[out_col] = matches
                df_out[master_col] = new_vals
            else:
                df_out[out_col] = "Column Missing"

        # ------------------------------------------------------------
        # Seniority parsing
        # ------------------------------------------------------------
        def parse_seniority(title):
            if not isinstance(title, str):
                return "Entry", "default: no seniority term found"
            t = title.lower().strip()
            if re.search(r"\bchief\b|\bcio\b|\bcto\b|\bceo\b|\bcfo\b|\bciso\b|\bcpo\b|\bcso\b|\bcoo\b|\bchro\b|\bpresident\b", t):
                return "C Suite", "keyword: c-level"
            if re.search(r"\bvice president\b|\bvp\b", t):
                return "VP", "keyword: vp"
            if re.search(r"\bhead\b", t):
                return "Head", "keyword: head"
            if re.search(r"\bdirector\b", t):
                return "Director", "keyword: director"
            if re.search(r"\bmanager\b|\bmgr\b", t):
                return "Manager", "keyword: manager/mgr"
            if re.search(r"\bsenior\b|\bsr\b|\blead\b|\bprincipal\b", t):
                return "Senior", "keyword: senior/lead"
            if re.search(r"\bintern\b|\btrainee\b|\bassistant\b|\bgraduate\b", t):
                return "Entry", "keyword: entry-level term"
            if re.search(r"\bengineer\b|\barchitect\b|\banalyst\b|\bdeveloper\b|\bconsultant\b|\bscientist\b|\btechnician\b|\bdesigner\b|\bassociate\b|\bcoordinator\b", t):
                return "Entry", "default: technical role"
            return "Entry", "default: none found"

        if "jobtitle" in df_master.columns:
            parsed = df_master["jobtitle"].apply(parse_seniority)
            df_out["Parsed_Seniority"] = parsed.apply(lambda x: x[0])
            df_out["Seniority_Logic"] = parsed.apply(lambda x: x[1])
        else:
            df_out["Parsed_Seniority"] = None
            df_out["Seniority_Logic"] = "jobtitle column not found"

        # ------------------------------------------------------------
        # Seniority match
        # ------------------------------------------------------------
        if "seniority" in df_picklist.columns:
            sen_set = set(normalize(df_picklist["seniority"]))
            df_out["Seniority_Match"] = df_out["Parsed_Seniority"].apply(
                lambda x: "Yes" if isinstance(x, str) and x.strip().lower() in sen_set else "No"
            )
        else:
            df_out["Seniority_Match"] = "Picklist Missing"

        # ------------------------------------------------------------
        # Write output file
        # ------------------------------------------------------------
        base_name = os.path.splitext(master_file.name)[0]
        out_file = f"{base_name} - Matched.xlsx"
        df_out.to_excel(out_file, index=False)

        wb = load_workbook(out_file)
        ws = wb.active
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        orig_cols = set(df_master.columns)
        new_cols = [c for c in df_out.columns if c not in orig_cols]

        for col_idx, col in enumerate(df_out.columns, start=1):
            if col in new_cols:
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).fill = yellow
                for row in range(2, ws.max_row + 1):
                    val = str(ws.cell(row=row, column=col_idx).value).strip().lower()
                    if val == "yes":
                        ws.cell(row=row, column=col_idx).fill = green
                    elif val == "no":
                        ws.cell(row=row, column=col_idx).fill = red

        for col_name, row in corrected_cells:
            if col_name in df_out.columns:
                idx = list(df_out.columns).index(col_name) + 1
                ws.cell(row=row, column=idx).fill = blue

        wb.save(out_file)
        progress(1.0, desc="‚úÖ Done! File ready for download.")
        return out_file

    except Exception as e:
        return f"‚ùå Error: {str(e)}"

# ------------------------------------------------------------
# Gradio Interface
# ------------------------------------------------------------
demo = gr.Interface(
    fn=run_matching,
    inputs=[
        gr.File(label="Upload MASTER Excel file (.xlsx)"),
        gr.File(label="Upload PICKLIST Excel file (.xlsx)")
    ],
    outputs=gr.File(label="Download Processed File"),
    title="üìä Master‚ÄìPicklist Matching Tool",
    description="Upload your MASTER and PICKLIST Excel files to perform automated matching and seniority parsing."
)

# ------------------------------------------------------------
# Launch for local or Railway deployment
# ------------------------------------------------------------
if __name__ == "__main__":
    demo.launch(
        server_name="0.0.0.0",
        server_port=int(os.environ.get("PORT", 7860))
    )
